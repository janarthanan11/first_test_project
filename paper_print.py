from openpyxl import Workbook
from openpyxl.styles import Alignment


paper_no = int(input("how many paper per square: "))
total_page = paper_no * 4

page_no_ltof = []

def valid_page(total_page):
    if total_page <= 4097:
        n = 1
        while True:
            if n < total_page:
                n = n * 2
            elif n == total_page:
                return True
            else:
                return False
    else:
        return False

def fold_val_fun(paper_no):
    root_val = paper_no
    fold_val = 1
    if paper_no != 1:
        while True:
            root_val = root_val // 2
            if root_val == 1:
                fold_val = fold_val + 1
                break
            else:
                fold_val = fold_val + 1
    return fold_val

def rcval(fold_val):
    if fold_val == 1:
        rc_val = {1: [1]}
    elif fold_val == 2:
        rc_val = {1: [1, 2]}
    else:
        i = 1
        a = 1
        b = 2
        k = 2
        for fold in range(3, fold_val+1, 2):
            rc_val = {}
            for j in range(k):
                if i%2:
                    rc_val[i] = [a, b]
                else:
                    rc_val[i] = [b, a]
                i = i + 1
                a = a + 2
                b = b + 2
            k = k * 2
            i = 1
            a = 1
            b = 2
    return rc_val

def b_rc_fun(fold_val):
    if fold_val == 1:
        rc_lis_val = ["1_1", "2_1"]
    elif fold_val == 2:
        rc_lis_val = ["1_1", "1_2", "2_1", "2_2"]
    else:
        rc_lis_val = []
        rc_dic = rcval(fold_val)
        rc_lis = ["1_1", "1_2", "2_1", "2_2"]
        
        for fold in range(3, fold_val + 1):
            rc_lis_val = []
            if fold % 2:
                for rc in rc_lis:
                    rc = rc.split("_")
                    rc_val = rc_dic[int(rc[0])]
                    for i in rc_val:
                        val = str(i) + "_" + rc[1]
                        rc_lis_val.append(val) 
            else:
                for rc in rc_lis:
                    rc = rc.split("_")
                    rc_val = rc_dic[int(rc[1])]
                    for i in rc_val:
                        val = rc[0] + "_" + str(i)
                        rc_lis_val.append(val) 
            rc_lis = rc_lis_val.copy()
            rc_lis.sort()    
    return rc_lis_val

def b_page_val_fun(all_page):
    len_val = len(all_page)
    if len_val == 1:
        page_val = [1, 4]
    elif len_val == 2:
        page_val = [4, 5, 1, 8]
    else:
        page_val = []
        page_nu = [4, 5, 1, 8]
        b = 2
        for k in range(3, len_val + 1):
            iv = 0
            rc_val = b_rc_fun(k)
            page_val = []
            page_val_ra = []
            new_page_val = []
            for i in page_nu:
                b_val = (all_page[b])[i-1]
                a_val = (all_page[b+1])[b_val-1]
                val = [b_val, a_val]
                page_val.extend(val)
            b = b + 1
            for rc in rc_val:
                page_val_ra.append(rc + "-" + str(page_val[iv])) 
                iv = iv + 1 
            page_val_ra.sort()
            for nu in page_val_ra:
                nu_sp = nu.split("-")
                new_page_val.append(int(nu_sp[1])) 
                
            page_nu = new_page_val
    return page_val

def f_rc_fun(fold_val):
    if fold_val == 1:
        rc_lis_val = ["1_1", "2_1"]
    elif fold_val == 2:
        rc_lis_val = ["1_1", "1_2", "2_1", "2_2"]
    else:
        rc_lis_val = []
        rc_dic = rcval(fold_val)
        rc_lis = ["1_1", "1_2", "2_1", "2_2"]
        
        for fold in range(3, fold_val + 1):
            rc_lis_val = []
            if fold % 2:
                for rc in rc_lis:
                    rc = rc.split("_")
                    rc_val = rc_dic[int(rc[0])]
                    for i in rc_val:
                        val = str(i) + "_" + rc[1]
                        rc_lis_val.append(val) 
            else:
                for rc in rc_lis:
                    rc = rc.split("_")
                    rc_val = rc_dic[int(rc[1])]
                    for i in rc_val:
                        val = rc[0] + "_" + str(i)
                        rc_lis_val.append(val) 
            rc_lis = rc_lis_val.copy()
            rc_lis.sort()    
    return rc_lis_val

def f_page_val_fun(all_page):
    len_val = len(all_page)
    if len_val == 1:
        page_val = [2, 3]
    elif len_val == 2:
        page_val = [6, 3, 7, 2]
    else:
        page_val = []
        page_nu = [6, 3, 7, 2]
        b = 2
        for k in range(3, len_val + 1):
            iv = 0
            rc_val = f_rc_fun(k)
            page_val = []
            page_val_ra = []
            new_page_val = []
            for i in page_nu:
                b_val = (all_page[b])[i-1]
                a_val = (all_page[b+1])[b_val-1]
                val = [b_val, a_val]
                page_val.extend(val)
            b = b + 1
            for rc in rc_val:
                page_val_ra.append(rc + "-" + str(page_val[iv])) 
                iv = iv + 1 
            page_val_ra.sort()
            for nu in page_val_ra:
                nu_sp = nu.split("-")
                new_page_val.append(int(nu_sp[1])) 
                
            page_nu = new_page_val
    return page_val

def print_excel(b_fin_page, f_fin_page, fold_val):
    workbook = Workbook()
    
    if fold_val%2:
        workbook.create_sheet("back_side", 0)
        sheet = workbook.active
        if fold_val == 1:
            sheet.cell(row=1, column=1).value = 1
            sheet.cell(row=1, column=1).alignment = Alignment(text_rotation=90, horizontal="right", vertical="top")
            sheet.row_dimensions[1].height = 100
            sheet.cell(row=2, column=1).value = 4
            sheet.cell(row=2, column=1).alignment = Alignment(text_rotation=90, horizontal="right", vertical="top")
            sheet.row_dimensions[2].height = 100
            workbook.save('page_no.xlsx')
            workbook.create_sheet("front_side", 1)
            sheet = workbook.get_sheet_by_name("front_side")
            sheet.cell(row=1, column=1).value = 2
            sheet.cell(row=1, column=1).alignment = Alignment(text_rotation=180, horizontal="left", vertical="bottom")
            sheet.row_dimensions[1].height = 100
            sheet.cell(row=2, column=1).value = 3
            sheet.cell(row=2, column=1).alignment = Alignment(text_rotation=180, horizontal="left", vertical="bottom")
            sheet.row_dimensions[2].height = 100
            workbook.save('page_no.xlsx')

        else:
            for i in b_fin_page:
                i = i.split("_")
                sheet.cell(row=int(i[0]), column=int(i[1])).value = int(i[2])
                sheet.row_dimensions[int(i[0])].height = 100
                if int(i[1])%2:
                    sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=90, horizontal="right", vertical="top")
                else:
                    sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=180, horizontal="left", vertical="bottom")
            workbook.save('page_no.xlsx')
        
            workbook.create_sheet("front_side", 1)
            sheet = workbook.get_sheet_by_name("front_side")

            for i in f_fin_page:
                i = i.split("_")
                sheet.cell(row=int(i[0]), column=int(i[1])).value = int(i[2])
                sheet.row_dimensions[int(i[0])].height = 100
                if int(i[1])%2:
                    sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=90, horizontal="right", vertical="top")
                else:
                    sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=180, horizontal="left", vertical="bottom")
            workbook.save('page_no.xlsx')
    else:
        workbook.create_sheet("back_side", 0)
        sheet = workbook.active
        for i in b_fin_page:
            i = i.split("_")
            sheet.cell(row=int(i[0]), column=int(i[1])).value = int(i[2])
            sheet.row_dimensions[int(i[0])].height = 60
            if int(i[0])%2:
                sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=0, horizontal="right", vertical="bottom")
            else:
                sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=180, horizontal="left", vertical="top")
        workbook.save('page_no.xlsx')
        
        workbook.create_sheet("front_side", 1)
        sheet = workbook.get_sheet_by_name("front_side")

        for i in f_fin_page:
            i = i.split("_")
            sheet.cell(row=int(i[0]), column=int(i[1])).value = int(i[2])
            sheet.row_dimensions[int(i[0])].height = 60
            if int(i[0])%2:
                sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=0, horizontal="right", vertical="bottom")
            else:
                sheet.cell(row=int(i[0]), column=int(i[1])).alignment = Alignment(text_rotation=180, horizontal="left", vertical="top")
        workbook.save('page_no.xlsx')
    return "completed!"


if valid_page(total_page):
    all_page = {}
    n = paper_no
    b_fin_page = []
    f_fin_page = []
    b = 0
    f = 0
    fold_val = fold_val_fun(paper_no)
    for i in range(fold_val):
        for page in range(total_page//n, 0, -1):
            page_no_ltof.append(page)
        page_no_ltof_val = page_no_ltof
        all_page[i+1] = page_no_ltof
        page_no_ltof = []
        n = n // 2
    b_rc = b_rc_fun(fold_val)
    b_page_val = b_page_val_fun(all_page)
    for i in b_rc:
        b_fin_page.append(i + "_" + str(b_page_val[b]))
        b = b + 1
    #print("back : ",b_fin_page)
    f_rc = f_rc_fun(fold_val)
    f_page_val = f_page_val_fun(all_page)
    for i in f_rc:
        f_fin_page.append(i + "_" + str(f_page_val[f]))
        f = f + 1
    #print("front : ",f_fin_page)
    print_excel(b_fin_page, f_fin_page, fold_val)


else:
    print("please enter less paper per square")